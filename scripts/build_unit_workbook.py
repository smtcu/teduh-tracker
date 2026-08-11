#!/usr/bin/env python3
"""Unit-type workbook: the working behind the Project Sales Insight tables.

One summary sheet of per-type counts per week, then one sheet per project listing
every unit with its block and classified type. This is where anyone can check a
figure back to individual unit numbers.

Usage: python3 scripts/build_unit_workbook.py <out.xlsx>
"""
import csv, os, sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter as L

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import unit_types as UT

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def F(**k):
    k.setdefault("size", 10)
    return Font(name="Arial", **k)


TH = Side(style="thin")
BORD = Border(left=TH, right=TH, top=TH, bottom=TH)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center")
HDR = PatternFill("solid", fgColor="D9E1F2")
NEWF = PatternFill("solid", fgColor="FCE4D6")


def cell(ws, r, c, v=None, font=None, align=CTR, fill=None, fmt=None):
    x = ws.cell(r, c, v)
    x.font = font or F(); x.alignment = align; x.border = BORD
    if fill: x.fill = fill
    if fmt: x.number_format = fmt
    return x


def read(path):
    p = os.path.join(ROOT, path)
    if not os.path.exists(p) or os.path.getsize(p) == 0:
        return []
    return list(csv.DictReader(open(p, newline="", encoding="utf-8")))


def main(out):
    spec = UT.config()
    weekly = read("data/teduh_unit_types_weekly.csv") or read("data/teduh_unit_types.csv")
    units = read("data/teduh_units.csv")

    wb = openpyxl.Workbook(); wb.calculation.fullCalcOnLoad = True
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "Unit sold by type — weekly record"
    ws["A1"].font = F(bold=True, size=12)
    ws["A2"] = "Generated " + datetime.now().strftime("%d/%m/%Y")
    ws["A2"].font = F(bold=True, color="C00000")

    r = 4
    for key, meta in spec.items():
        types = [t["key"] for t in meta["types"]]
        totals = {t["key"]: t["total"] for t in meta["types"]}
        rows = [x for x in weekly if x.get("project_key") == key]
        weeks = sorted({x["week"] for x in rows})
        got = {(x["week"], x["unit_type"]): int(float(x["sold"])) for x in rows if x.get("sold") not in ("", None)}

        cell(ws, r, 1, meta["label"], F(bold=True, size=11), align=LFT)
        r += 1
        cell(ws, r, 1, "WEEK", F(bold=True), fill=HDR)
        for i, t in enumerate(types):
            cell(ws, r, 2 + i, f"{t}\n({[x['size'] for x in meta['types']][i]})", F(bold=True), fill=HDR)
        cell(ws, r, 2 + len(types), "TOTAL", F(bold=True), fill=HDR)
        cell(ws, r, 3 + len(types), "%", F(bold=True), fill=HDR)
        head = r
        r += 1
        cell(ws, r, 1, "TOTAL UNITS", F(bold=True), align=LFT, fill=HDR)
        for i, t in enumerate(types):
            cell(ws, r, 2 + i, totals[t], F(bold=True), fmt="#,##0", fill=HDR)
        cell(ws, r, 2 + len(types), sum(totals.values()), F(bold=True), fmt="#,##0", fill=HDR)
        base = r
        r += 1
        for wi, w in enumerate(weeks):
            last = wi == len(weeks) - 1
            fill = NEWF if last else None
            cell(ws, r, 1, datetime.strptime(w, "%Y-%m-%d"), fmt="DD/MM/YYYY", align=LFT, fill=fill)
            for i, t in enumerate(types):
                cell(ws, r, 2 + i, got.get((w, t)), fmt="#,##0", fill=fill)
            c0, c1 = L(2), L(1 + len(types))
            cell(ws, r, 2 + len(types), f"=SUM({c0}{r}:{c1}{r})", fmt="#,##0", fill=fill)
            cell(ws, r, 3 + len(types),
                 f"={L(2 + len(types))}{r}/{L(2 + len(types))}{base}", fmt="0.0%", fill=fill)
            r += 1
        ws.column_dimensions["A"].width = 22
        for i in range(len(types) + 2):
            ws.column_dimensions[L(2 + i)].width = 11
        r += 2

    # ---- one sheet per project, every unit listed ----
    for key, meta in spec.items():
        mine = [u for u in units if u.get("project_key") == key]
        if not mine:
            continue
        latest = max(u["week"] for u in mine)
        mine = sorted((u for u in mine if u["week"] == latest), key=lambda u: u["unit"])
        sh = wb.create_sheet(meta["label"][:31])
        sh["A1"] = f"{meta['label']} — unit list as at {latest}"
        sh["A1"].font = F(bold=True, size=12)
        for i, h in enumerate(["No.", "Unit Number", "Block", "Unit Type", "Sold"], 1):
            cell(sh, 3, i, h, F(bold=True), fill=HDR)
        for i, u in enumerate(mine, 1):
            cell(sh, 3 + i, 1, i)
            cell(sh, 3 + i, 2, u["unit"], align=LFT)
            cell(sh, 3 + i, 3, u["block"])
            cell(sh, 3 + i, 4, u["unit_type"])
            cell(sh, 3 + i, 5, "Yes" if u["sold"] in ("1", 1) else "")
        for c, w in [("A", 6), ("B", 16), ("C", 9), ("D", 11), ("E", 8)]:
            sh.column_dimensions[c].width = w
        sh.freeze_panes = "A4"

    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
    wb.save(out)
    print(f"{out}  ({len(spec)} projects, {len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main(sys.argv[1])
