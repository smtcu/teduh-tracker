#!/usr/bin/env python3
"""Rebuild both TEDUH weekly Excel trackers from projects.csv + teduh_history.csv.

Usage: python3 build_trackers.py <projects.csv> <teduh_history.csv> <outdir> [report_date YYYY-MM-DD]

Layout, with APDL DATE and TOTAL UNIT ahead of the code and developer:
  seputeh  : NO | PROJECT | APDL DATE | TOTAL UNIT | CODE | DEVELOPER | (NEW SALES,TOTAL SOLD,%) x weeks | Remarks
  status13 : NO | PROJECT | APDL DATE | TOTAL UNIT | CODE | DEVELOPER | (SOLD,%) | (NEW SALES,TOTAL SOLD,%) x weeks
  grouped  : NO | PROJECT | APDL DATE | TOTAL UNIT | DEVELOPER | (NEW SALES,TOTAL SOLD,TOTAL %) x weeks | NOTES
APDL DATE is the earliest advertising-permit start, stored per project in
projects.csv rather than fetched, because it never changes once issued.
It replaced the Launched date column on 01 Sep 2026.
NEW SALES and % are live Excel formulas, never hardcoded results.
"""
import csv, sys, os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter as L

def F(**k):
    k.setdefault('size', 10)
    return Font(name='Arial', **k)

TH = Side(style='thin'); BORD = Border(left=TH, right=TH, top=TH, bottom=TH)
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
HDR = PatternFill('solid', fgColor='D9E1F2')
NEWFILL = PatternFill('solid', fgColor='FFF2CC')   # highlights the newest week

AREA_TRACKERS = ('seputeh', 'status13', 'johor', 'ukay')


def developer_trackers(projects):
    """[(tracker key, sheet name)] for every non-area tracker, A-Z by label.

    Anything in projects.csv that is not one of the four area sheets is a
    competitor watch. Read from the data rather than listed here, so adding a
    fourteenth company means adding rows to projects.csv and nothing else.
    """
    labels = {}
    for p in projects:
        key = (p.get('tracker') or '').strip()
        if key and key not in AREA_TRACKERS and key not in labels:
            labels[key] = (p.get('tracker_label') or '').strip() or key.title()
    return sorted(labels.items(), key=lambda kv: kv[1].lower())


def apdl_date(p):
    """Earliest advertising-permit start, as stored in projects.csv."""
    s = (p.get('apdl') or '').strip()
    return datetime.strptime(s, '%Y-%m-%d') if s else None

def cell(ws, r, c, v=None, font=None, align=CTR, fill=None, fmt=None):
    x = ws.cell(r, c, v)
    x.font = font or F(); x.alignment = align; x.border = BORD
    if fill: x.fill = fill
    if fmt: x.number_format = fmt
    return x


def remark_for(project, generated):
    """Same rule as the website: `remarks` overrides outright, `note_prefix` is a
    standing caveat that keeps the live block numbers after it. Kept in step with
    build_dashboard.remark_for so the Excel and the site never disagree."""
    override = (project.get('remarks') or '').strip()
    if override:
        return override
    caveat = (project.get('note_prefix') or '').strip()
    return ' '.join(part for part in (caveat, generated) if part)

def seed_missing(projects, order, lookup, daily_csv):
    """Give a tracker with no weekly record one column, taken from today's reading.

    A competitor tracker added midweek has daily readings but nothing yet in
    teduh_history.csv, so its sheet would come out as a list of projects with no
    figures at all. Seeding it means the sheet reads properly the day it is
    created, and the first real Friday run replaces the seeded column with a
    recorded one. Trackers that already have weekly history are never touched.

    Same rule as build_dashboard.build_payload, including dating the column at
    the newest Friday rather than at today, so the seeded column lines up with
    the other sheets instead of inventing a date of its own.
    """
    if not daily_csv or not os.path.exists(daily_csv):
        return set()
    daily = list(csv.DictReader(open(daily_csv, encoding='utf-8')))
    if not daily:
        return set()
    latest_day = max(r['week'] for r in daily)
    dated_at = max((wk for snaps in order.values() for _, wk in snaps), default=latest_day)

    missing = {(p.get('tracker') or '').strip() for p in projects} - set(order) - {''}
    seeded = set()
    for t in sorted(missing):
        rows = [r for r in daily if r.get('tracker') == t and r['week'] == latest_day]
        if not rows:
            continue
        snap = (1, dated_at)
        order[t] = [snap]
        for r in rows:
            try:
                lookup[(t, snap, r['code'])] = int(float(r['total_sold']))
            except (TypeError, ValueError):
                continue
        seeded.add(t)
    return seeded


def load(projects_csv, history_csv, daily_csv=None):
    """Return (projects, {tracker: [snapshot_keys...]}, lookup, seeded_trackers).

    A snapshot is one weekly column. It is keyed by (seq, week) rather than the
    date alone, because the tracker legitimately contains two columns bearing the
    same date -- deduping on date would silently drop a week of history.
    Rows appended by the scraper have a blank seq; they are assigned the next
    sequence number for their tracker, in date order.
    """
    projects = list(csv.DictReader(open(projects_csv, encoding='utf-8')))
    hist = list(csv.DictReader(open(history_csv, encoding='utf-8')))

    maxseq = {}
    for h in hist:
        if str(h.get('seq', '')).strip().isdigit():
            t = h['tracker']
            maxseq[t] = max(maxseq.get(t, 0), int(h['seq']))
    assigned = {}
    for h in sorted(hist, key=lambda x: (x['tracker'], x['week'])):
        if not str(h.get('seq', '')).strip().isdigit():
            k = (h['tracker'], h['week'])
            if k not in assigned:
                maxseq[h['tracker']] = maxseq.get(h['tracker'], 0) + 1
                assigned[k] = maxseq[h['tracker']]

    order, lookup = {}, {}
    for h in hist:
        t = h['tracker']
        seq = int(h['seq']) if str(h.get('seq', '')).strip().isdigit() else assigned[(t, h['week'])]
        snap = (seq, h['week'])
        order.setdefault(t, set()).add(snap)
        v = h['total_sold']
        try: v = int(float(v)) if v not in ('', 'None', None) else None
        except (TypeError, ValueError): v = None
        lookup[(t, snap, h['code'])] = v
    order = {t: sorted(v) for t, v in order.items()}
    seeded = seed_missing(projects, order, lookup, daily_csv)
    return projects, order, lookup, seeded

def build_seputeh(projects, weeks, lookup, report_date, path):
    wb = openpyxl.Workbook(); wb.calculation.fullCalcOnLoad = True; ws = wb.active; ws.title = report_date.strftime('%Y%m%d')
    ws['A1'] = 'Teduh Weekly Update: Seputeh Hills Competitor Studies'; ws['A1'].font = F(bold=True, size=12)
    ws['A2'] = 'Report as at ' + report_date.strftime('%d/%m/%Y'); ws['A2'].font = F(bold=True, color='FF0000')
    for i, h in enumerate(['NO','PROJECT','APDL \nDATE','TOTAL \nUNIT','PROJECT \nCODE','DEVELOPER'], 1):
        cell(ws, 4, i, h, F(bold=True), fill=HDR)
    col = 7
    for i, (_, wk) in enumerate(weeks):
        last = (i == len(weeks) - 1)
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+2)
        cell(ws, 3, col, datetime.strptime(wk, '%Y-%m-%d'), F(bold=True), fmt='DD.MM.YYYY', fill=NEWFILL if last else None)
        for j, h in enumerate(['NEW SALES','TOTAL SOLD','%']):
            cell(ws, 4, col+j, h, F(bold=True), fill=NEWFILL if last else HDR)
        col += 3
    rem = col
    cell(ws, 4, rem, 'Remarks', F(bold=True), fill=HDR)

    rows = sorted((p for p in projects if p['tracker'] == 'seputeh'),
                  key=lambda x: (x.get('pin', '').strip().lower() not in ('yes', 'y', '1', 'true')))
    for r, p in enumerate(rows, 5):
        cell(ws, r, 1, int(p['no']))
        cell(ws, r, 2, p['project'], align=LFT)
        cell(ws, r, 3, apdl_date(p), fmt='DD/MM/YYYY')
        cell(ws, r, 4, int(p['total_units']), fmt='#,##0')
        cell(ws, r, 5, p['code'])
        cell(ws, r, 6, p['developer'], align=LFT)
        col, prev = 7, None
        for i, snap in enumerate(weeks):
            v = lookup.get(('seputeh', snap, p['code']))
            fill = NEWFILL if i == len(weeks) - 1 else None
            n, t, pc = cell(ws, r, col, fmt='#,##0', fill=fill), cell(ws, r, col+1, fmt='#,##0', fill=fill), cell(ws, r, col+2, fmt='0.0%', fill=fill)
            if v is not None:
                t.value = v
                n.value = f'={L(col+1)}{r}-{prev}{r}' if prev else (int(p['first_new']) if str(p['first_new']).strip() else None)
                pc.value = f'={L(col+1)}{r}/$D${r}'
                prev = L(col+1)
            col += 3
        cell(ws, r, rem, p['remarks'], align=LFT)

    for c, w in [('A',4),('B',26),('C',12),('D',8),('E',10),('F',28)]:
        ws.column_dimensions[c].width = w
    for c in range(7, rem): ws.column_dimensions[L(c)].width = 9
    ws.column_dimensions[L(rem)].width = 42
    ws.freeze_panes = 'G5'; ws.row_dimensions[4].height = 30
    wb.save(path); return path

def build_status13(projects, weeks, lookup, report_date, path):
    wb = openpyxl.Workbook(); wb.calculation.fullCalcOnLoad = True; ws = wb.active; ws.title = report_date.strftime('%d%m%Y')
    ws['A2'] = 'Report as at ' + report_date.strftime('%d/%m/%Y'); ws['A2'].font = F(bold=True, color='FF0000')
    ws['A3'] = 'TEDUH WEEKLY UPDATE '; ws['A3'].font = F(bold=True, size=12)
    for i, h in enumerate(['NO','PROJECT','APDL \nDATE','TOTAL \nUNIT','PROJECT \nCODE','DEVELOPER'], 1):
        cell(ws, 4, i, h, F(bold=True), fill=HDR)
    ws.merge_cells('G3:H3')
    cell(ws, 3, 7, datetime.strptime(weeks[0][1], '%Y-%m-%d'), F(bold=True), fmt='DD/MM/YYYY')
    cell(ws, 4, 7, 'SOLD', F(bold=True), fill=HDR); cell(ws, 4, 8, '%', F(bold=True), fill=HDR)
    col = 9
    for i, (_, wk) in enumerate(weeks[1:], 1):
        last = (i == len(weeks) - 1)
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+2)
        cell(ws, 3, col, datetime.strptime(wk, '%Y-%m-%d'), F(bold=True), fmt='DD/MM/YYYY', fill=NEWFILL if last else None)
        for j, h in enumerate(['NEW SALES ','TOTAL SOLD','%']):
            cell(ws, 4, col+j, h, F(bold=True), fill=NEWFILL if last else HDR)
        col += 3
    end = col

    rows = sorted((p for p in projects if p['tracker'] == 'status13'),
                  key=lambda x: (x.get('pin', '').strip().lower() not in ('yes', 'y', '1', 'true')))
    for r, p in enumerate(rows, 5):
        key = p['code'] if p['code'] else f'NOCODE-R{r}'
        cell(ws, r, 1, int(p['no']))
        cell(ws, r, 2, p['project'], align=LFT)
        cell(ws, r, 3, apdl_date(p), fmt='DD/MM/YYYY')
        cell(ws, r, 4, int(p['total_units']), fmt='#,##0')
        cell(ws, r, 5, p['code'])
        cell(ws, r, 6, p['developer'], align=LFT)
        v0 = lookup.get(('status13', weeks[0], key))
        cell(ws, r, 7, v0, fmt='#,##0')
        cell(ws, r, 8, f'=G{r}/$D${r}' if v0 is not None else None, fmt='0.0%')
        col, prev = 9, ('G' if v0 is not None else None)
        for i, snap in enumerate(weeks[1:], 1):
            v = lookup.get(('status13', snap, key))
            fill = NEWFILL if i == len(weeks) - 1 else None
            n, t, pc = cell(ws, r, col, fmt='#,##0', fill=fill), cell(ws, r, col+1, fmt='#,##0', fill=fill), cell(ws, r, col+2, fmt='0.0%', fill=fill)
            if v is not None:
                t.value = v
                if prev: n.value = f'={L(col+1)}{r}-{prev}{r}'
                pc.value = f'={L(col+1)}{r}/$D${r}'
                prev = L(col+1)
            col += 3

    for c, w in [('A',4),('B',26),('C',12),('D',8),('E',10),('F',30)]:
        ws.column_dimensions[c].width = w
    for c in range(7, end): ws.column_dimensions[L(c)].width = 9
    ws.freeze_panes = 'G5'; ws.row_dimensions[4].height = 30
    wb.save(path); return path

def fill_grouped(ws, projects, weeks, lookup, notes, report_date, tracker='johor',
                 heading='WEEKLY TEDUH SALES REPORT (JB PROJECTS)', show_code=False,
                 seeded=False):
    """Fill one worksheet: every week ever recorded, optional group headings, NOTES.

    Shared by Johor, Ukay and all thirteen developer sheets. They differ only in
    which rows they take, the heading, whether `group` is set -- Ukay leaves it
    blank, so it renders as one flat list with no section rows -- and whether the
    PROJECT CODE column is shown. The developer sheets show it: the whole point of
    those sheets is which TEDUH code a competitor's project turned out to be.

    Split out of build_grouped so a workbook can hold many of these sheets.
    """
    ws['A2'] = 'Report as at ' + report_date.strftime('%d/%m/%Y'); ws['A2'].font = F(bold=True, color='FF0000')
    ws['A3'] = heading; ws['A3'].font = F(bold=True, size=12)
    if seeded:
        ws['A1'] = ('New tracker: the single column below is today\'s reading standing in '
                    'for a weekly record, so it has no NEW SALES figure yet.')
        ws['A1'].font = F(bold=True, color='C00000', size=9)
    heads = ['NO', 'PROJECT', 'APDL \nDATE', 'TOTAL \nUNIT']
    if show_code:
        heads.append('PROJECT \nCODE')
    heads.append('DEVELOPER')
    for i, h in enumerate(heads, 1):
        cell(ws, 4, i, h, F(bold=True), fill=HDR)
    col = len(heads) + 1
    first_week_col = col
    for i, snap in enumerate(weeks):
        last = (i == len(weeks) - 1)
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
        cell(ws, 3, col, datetime.strptime(snap[1], '%Y-%m-%d'), F(bold=True), fmt='DD/MM/YYYY',
             fill=NEWFILL if last else None)
        for j, h in enumerate(['NEW SALES', 'TOTAL SOLD', 'TOTAL %']):
            cell(ws, 4, col + j, h, F(bold=True), fill=NEWFILL if last else HDR)
        col += 3
    rem_col = col
    cell(ws, 4, rem_col, 'NOTES', F(bold=True), fill=HDR)

    rows = [p for p in projects if p['tracker'] == tracker]
    r = 5
    group = None
    for p in rows:
        if p.get('group') and p['group'] != group:
            group = p['group']
            cell(ws, r, 1, group, F(bold=True), align=LFT, fill=HDR)
            for c in range(2, rem_col + 1):
                cell(ws, r, c, None, fill=HDR)
            r += 1
        key = (p.get('code') or '').split(',')[0].strip() or f"NOCODE-{p['project']}"
        units = int(p['total_units']) if str(p.get('total_units', '')).strip().isdigit() else 0
        cell(ws, r, 1, int(p['no']))
        cell(ws, r, 2, p['project'], align=LFT)
        cell(ws, r, 3, apdl_date(p), fmt='DD/MM/YYYY')
        cell(ws, r, 4, units, fmt='#,##0')
        if show_code:
            cell(ws, r, 5, p.get('code') or '')
        cell(ws, r, len(heads), p['developer'], align=LFT)
        col, prev = first_week_col, None
        for i, snap in enumerate(weeks):
            v = lookup.get((tracker, snap, key))
            fill = NEWFILL if i == len(weeks) - 1 else None
            n = cell(ws, r, col, fmt='#,##0', fill=fill)
            t = cell(ws, r, col + 1, fmt='#,##0', fill=fill)
            pc = cell(ws, r, col + 2, fmt='0.0%', fill=fill)
            if v is not None:
                t.value = v
                if prev:
                    n.value = f'={L(col + 1)}{r}-{prev}{r}'
                pc.value = f'={L(col + 1)}{r}/$D${r}'
                prev = L(col + 1)
            col += 3
        cell(ws, r, rem_col, remark_for(p, notes.get(key, '')), align=LFT)
        r += 1

    # C is 13, not 12: at 12 Excel is a character short of DD/MM/YYYY and renders
    # the whole APDL column as #####.
    widths = [('A', 4), ('B', 26), ('C', 13), ('D', 9)]
    widths += [('E', 12), ('F', 26)] if show_code else [('E', 22)]
    for c, w in widths:
        ws.column_dimensions[c].width = w
    for c in range(first_week_col, rem_col):
        ws.column_dimensions[L(c)].width = 9
    ws.column_dimensions[L(rem_col)].width = 46
    ws.freeze_panes = L(first_week_col) + '5'
    ws.row_dimensions[4].height = 30
    return ws


def build_grouped(projects, weeks, lookup, notes, report_date, path, tracker='johor',
                  heading='WEEKLY TEDUH SALES REPORT (JB PROJECTS)'):
    """One tracker, one workbook -- the Johor and Ukay files."""
    wb = openpyxl.Workbook(); wb.calculation.fullCalcOnLoad = True
    ws = wb.active; ws.title = report_date.strftime('%d%m%Y')
    fill_grouped(ws, projects, weeks, lookup, notes, report_date, tracker, heading)
    wb.save(path); return path


def build_developers(projects, order, lookup, notes, report_date, path, seeded=()):
    """All the competitor trackers in one workbook, one sheet per company.

    One file rather than thirteen: thirteen more .xlsx in docs/downloads would
    bury the four area trackers people actually open, and the Friday cleanup step
    globs that directory by name.
    """
    devs = [(k, name) for k, name in developer_trackers(projects) if k in order]
    if not devs:
        return None
    wb = openpyxl.Workbook(); wb.calculation.fullCalcOnLoad = True
    wb.remove(wb.active)
    for key, name in devs:
        ws = wb.create_sheet(name[:31])
        fill_grouped(ws, projects, order[key], lookup, notes, report_date,
                     tracker=key, heading=f'WEEKLY TEDUH SALES REPORT ({name.upper()})',
                     show_code=True, seeded=key in seeded)
    wb.save(path)
    return path


if __name__ == '__main__':
    flags = [a for a in sys.argv[1:] if a.startswith('--')]
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    # --seed-from=data/teduh_daily.csv gives a brand-new tracker one column from
    # today's reading instead of an empty sheet. Optional and explicit: without
    # it this script reads nothing but the history file, as it always has.
    seed_from = next((f.split('=', 1)[1] for f in flags if f.startswith('--seed-from=')), None)
    pcsv, hcsv, outdir = args[0], args[1], args[2]
    rd = datetime.strptime(args[3], '%Y-%m-%d') if len(args) > 3 else datetime.now()
    os.makedirs(outdir, exist_ok=True)
    projects, order, lookup, seeded = load(pcsv, hcsv, seed_from)
    notes = {}
    for row in csv.DictReader(open(hcsv, encoding='utf-8')):
        n = (row.get('block_note') or '').strip()
        if n:
            notes[row.get('code', '')] = n
    p1 = build_seputeh(projects, order['seputeh'], lookup, rd,
                       os.path.join(outdir, rd.strftime('%Y%m%d') + '_Seputeh Hills_Teduh Weekly Update.xlsx'))
    p2 = build_status13(projects, order['status13'], lookup, rd,
                        os.path.join(outdir, rd.strftime('%d%m%Y') + 'Tduh Developer Project Sales Status 13.xlsx'))
    print(p1); print(p2)
    if 'johor' in order:
        p3 = build_grouped(projects, order['johor'], lookup, notes, rd,
                           os.path.join(outdir, rd.strftime('%Y%m%d') + '_Johor_Teduh_Weekly_Update.xlsx'))
        print(p3)
    if 'ukay' in order:
        p4 = build_grouped(projects, order['ukay'], lookup, notes, rd,
                           os.path.join(outdir, rd.strftime('%Y%m%d') + '_Ukay_Teduh_Weekly_Update.xlsx'),
                           tracker='ukay', heading='WEEKLY TEDUH SALES REPORT (UKAY PROJECTS)')
        print(p4)
    # Fixed filename, and deliberately "Competitor" rather than "Developer": the
    # Friday cleanup step globs *Developer*.xlsx to find the dated Klang Valley
    # workbook, and would copy this straight over it.
    p5 = build_developers(projects, order, lookup, notes, rd,
                          os.path.join(outdir, 'TEDUH_Competitor_Trackers.xlsx'),
                          seeded=seeded)
    if p5:
        print(p5)
