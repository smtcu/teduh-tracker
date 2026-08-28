#!/usr/bin/env python3
"""Turn the developer index into a workbook for cross-checking by hand.

Reads data/developers.csv (from discover_developers.py) and companies.csv, and
writes one sheet listing every developer found plus one candidate sheet per
company. Candidates are matched on company name, email domain and address
separately, and the sheet says which signal fired -- because no single signal
is reliable. Chin Hin registers companies with unrelated names and a Fiamma
email, and is only identifiable by "Menara Chin Hin" in the address; TS Law is
only identifiable by its email; some developers use a personal gmail and match
nothing at all.

  python scripts/build_developer_workbook.py data/TEDUH_Company_Index.xlsx

Writes to data/, not docs/downloads/. Two reasons: the website must not change
while the new trackers are still being designed, and the Friday cleanup globs
docs/downloads for "*Developer*.xlsx" and copies the newest match onto the
Klang Valley tracker -- a file named TEDUH_Developers.xlsx there would quietly
overwrite it.
"""
import csv, os, re, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FONT = "Arial"
HEAD = PatternFill("solid", fgColor="2A78D6")
KNOWN = PatternFill("solid", fgColor="FFF2CC")      # a code you already confirmed

COLUMNS = [("kod_pemaju", "DEVELOPER CODE", 15), ("nama_pemaju", "COMPANY NAME", 42),
           ("emel", "EMAIL", 30), ("telefon", "PHONE", 15),
           ("alamat_perniagaan", "BUSINESS ADDRESS", 52),
           ("alamat_daftar", "REGISTERED ADDRESS", 52),
           ("status_pemaju", "STATUS", 12), ("bilangan_projek", "PROJECTS", 10),
           ("negeri", "STATE", 18), ("daerah", "DISTRICT", 18),
           ("projek_nama", "FIRST PROJECT", 34)]


def read(path):
    if not os.path.exists(path):
        sys.exit(f"missing {path} — run discover_developers.py first")
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def parts(value):
    return [p.strip().lower() for p in (value or "").split("|") if p.strip()]


def matches(dev, rule):
    """Which of the three signals identify this developer as the company's."""
    hits = []
    name = (dev.get("nama_pemaju") or "").lower()
    mail = (dev.get("emel") or "").lower()
    addr = ((dev.get("alamat_perniagaan") or "") + " " +
            (dev.get("alamat_daftar") or "")).lower()
    if any(p in name for p in parts(rule.get("match_name"))):
        hits.append("name")
    if any(p in mail for p in parts(rule.get("match_email"))):
        hits.append("email")
    if any(p in addr for p in parts(rule.get("match_address"))):
        hits.append("address")
    return hits


def sheet(wb, title, rows, extra=None, known=()):
    ws = wb.create_sheet(title[:31])
    headers = [h for _, h, _ in COLUMNS] + ([extra] if extra else [])
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEAD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for c, (_, _, w) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    if extra:
        ws.column_dimensions[get_column_letter(len(COLUMNS) + 1)].width = 40

    for r, row in enumerate(rows, 2):
        for c, (key, _, _) in enumerate(COLUMNS, 1):
            v = row.get(key, "")
            if key in ("kod_pemaju", "bilangan_projek") and str(v).strip().isdigit():
                v = int(v)
            cell = ws.cell(r, c, v)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (5, 6)))
        if extra:
            cell = ws.cell(r, len(COLUMNS) + 1, row.get("_confidence", row.get("_matched", "")))
            cell.font = Font(name=FONT, size=10)
            v = str(cell.value)
            if v.startswith("CHECK"):
                cell.fill = PatternFill("solid", fgColor="FDE2D5")
            elif v.startswith("CONFIRMED"):
                cell.fill = PatternFill("solid", fgColor="D6EAD6")
        if str(row.get("kod_pemaju", "")) in known:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = KNOWN

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(rows) + 1)}"
    ws.row_dimensions[1].height = 28
    return ws


def legend(wb, devs, companies, found):
    ws = wb.create_sheet("How to use", 0)
    lines = [
        ("TEDUH developer index", True),
        ("", False),
        (f"{found} developers found across {len(devs)} codes probed.", False),
        ("", False),
        ("A TEDUH project code is the developer's code plus a sequence number, so 31332-1", False),
        ("is project 1 of developer 31332. Once a company's developer codes are known, every", False),
        ("project it owns follows automatically — including phases filed under names that do", False),
        ("not match how the project is advertised.", False),
        ("", False),
        ("Each company sheet lists candidates matched on THREE separate signals, and the", False),
        ("last column says which one fired. No single signal is reliable:", False),
        ("", False),
        ("   name      Exsim registers companies called EXSIM SOMETHING — works.", False),
        ("   email     TS Law registers PLAZA SENTOSA PROPERTIES but uses tslawland.com.", False),
        ("   address   Chin Hin registers AFFLUENT CRAFTS with a fiamma.com.my email;", False),
        ("             only 'Menara Chin Hin' in the address gives it away.", False),
        ("", False),
        ("Some developers use a personal gmail and a generic name, and will match nothing.", False),
        ("Those have to be found from a project you already know belongs to the company.", False),
        ("", False),
        ("Shaded rows are codes already confirmed in companies.csv.", False),
        ("", False),
        ("Check the candidates, then put the confirmed codes into companies.csv.", False),
    ]
    for r, (text, bold) in enumerate(lines, 1):
        c = ws.cell(r, 1, text)
        c.font = Font(name=FONT, size=13 if bold else 10, bold=bold)
    ws.column_dimensions["A"].width = 100
    return ws


def squash(s):
    """Loose key for comparing project names across sources."""
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


# Words that carry no identity: they appear in half the names on TEDUH.
NOISE = {"residensi", "residence", "residences", "apartment", "apartments",
         "pangsapuri", "kondominium", "condominium", "taman", "bandar", "fasa",
         "phase", "the", "at", "by", "and", "sdn", "bhd", "suites", "suite",
         "tower", "towers", "block", "blok", "development", "project"}


def tokens(name, drop=()):
    """Identity-bearing words in a project name.

    TEDUH reorders and prefixes names -- your "Park One Residence" is filed as
    "Residensi Park One" -- so comparing whole strings misses real matches.
    Place names are dropped because they are the separate `place` signal; left
    in, "Taman Melawati Fasa 3" would masquerade as "16 Quartz @ Taman Melawati".

    Letters and digits are split apart, because TEDUH glues them together:
    "Nadayu63 @ Melawati" has to reach your "Nadayu 63", and "20trees" your
    "20Trees". Splitting does not blur 19Trees into 20Trees -- they still share
    only the word "trees", which is not enough on its own.
    """
    words = re.findall(r"[a-z]+|[0-9]+", str(name or "").lower())
    return {w for w in words if w not in NOISE and w not in drop and len(w) > 1}


def name_match(a, b, drop=()):
    """True when two project names plausibly describe the same development.

    Compares identity words only. A plain substring test is far too generous
    here: "DaMai" sits inside Taman Damai Impian, Damai Suria, Bandar Damai
    Perdana and 167 other unrelated Malaysian projects.

    So a single shared word is only accepted when it is long enough to be a
    real name and is essentially the whole of the shorter title -- "Wellnessa"
    identifies Trinity Wellnessa, "Hijauan" does not identify Sierra Hijauan,
    because half the country has Hijauan in it.

    Returns "strong", "weak" or None. Weak means the names agree on a single
    short word -- your "DaMai" against a project actually called "Taman Damai".
    That is real evidence but not proof, so a weak match has to be backed by a
    second signal before it earns a place on the sheet.
    """
    ta, tb = tokens(a, drop), tokens(b, drop)
    if not ta or not tb:
        return None
    if squash(a) == squash(b):          # the same name, written the same way
        return "strong"
    shared = ta & tb
    if not shared:
        return None
    if len(shared) >= 2:
        return "strong"
    word = next(iter(shared))
    shorter = ta if len(ta) <= len(tb) else tb
    if shared != shorter:               # the shorter name says more than this
        return None
    if len(word) >= 8:                  # long enough to be a real project name
        return "strong"
    return "weak" if ta == tb else None


def build_areas(wb, found, ppath):
    """One sheet per area in areas.csv, listing projects rather than developers.

    An area tracker asks "what is being built here", so it matches at project
    level and on four independent signals, because no single one is reliable:

      project    the name you already track, matched loosely against TEDUH's
      place      an Ukay-area place name in the project name or the address
      developer  a developer you know builds in the area
      district   the wide sweep, Gombak and Hulu Langat

    District alone is deliberately NOT enough to list a project: Gombak also
    covers Selayang and Kundang, and the Melawati side sits in the Kuala Lumpur
    district, so district on its own would be both noisy and incomplete. It is
    still reported, and the sweep sheet shows everything in those districts so
    nothing is silently dropped.

    Where you already know the advertised unit count, it is put beside TEDUH's
    and compared -- that is what confirms identity when the names diverge, the
    same way Gen Sphere's 996 tied to Residensi Gen Sfera.
    """
    apath = os.path.join(ROOT, "areas.csv")
    if not os.path.exists(apath) or not os.path.exists(ppath):
        return
    areas = read(apath)
    projects = read(ppath)
    known = read(os.path.join(ROOT, "area_projects.csv")) \
        if os.path.exists(os.path.join(ROOT, "area_projects.csv")) else []
    by_dev = {d["kod_pemaju"]: d for d in found}

    for rule in areas:
        area = rule["area"]
        places = parts(rule.get("place_keywords"))
        devs = parts(rule.get("developer_keywords"))
        districts = {p.lower() for p in parts(rule.get("districts"))}
        states = {p.lower() for p in parts(rule.get("states"))}
        mine = [k for k in known if k.get("area", "").lower() == area.lower()]
        names = {squash(k["project"]): k for k in mine if k.get("project")}

        # An exact unit count is strong evidence inside a small area, and it is
        # the only thing that finds a project TEDUH has renamed beyond
        # recognition -- Park One Residence is filed as "Residensi Laman Satu
        # Melawati", but both are 234 units. Only counts that identify exactly
        # one of your projects are used; a figure shared by two of them proves
        # nothing.
        by_units = {}
        for k in mine:
            v = str(k.get("advertised_units", "")).strip()
            if v.isdigit():
                by_units.setdefault(int(v), []).append(k)
        by_units = {u: v[0] for u, v in by_units.items() if len(v) == 1}

        rows, sweep = [], []
        for p in projects:
            dev = by_dev.get(p.get("kod_pemaju"), {})
            pname = (p.get("projek_nama") or "")
            dname = (dev.get("nama_pemaju") or "")
            addr = ((dev.get("alamat_perniagaan") or "") + " " +
                    (dev.get("alamat_daftar") or "")).lower()
            hay = (pname + " " + addr).lower()
            dist = (p.get("daerah") or "").strip().lower()
            neg = (p.get("negeri") or "").strip().lower()
            if dist in ("-", ""):           # TEDUH leaves these blank on older records
                dist = ""
            if neg in ("-", ""):
                neg = ""

            # Geography is a veto, not a signal. A project whose state or
            # district is known and wrong is not in this area, however much its
            # name looks right -- PKNS builds a "Melawati" in Kuala Selangor,
            # and Sime Darby's Saujana Impian sits in Hulu Langat, neither of
            # which is Ukay. Blank means unknown, so it is allowed through.
            if states and neg and neg not in states:
                continue
            if districts and dist and dist not in districts:
                continue

            drop = {w for k in places for w in re.findall(r"[a-z0-9]+", k)}
            hits, matched_to, grade = [], None, None
            for rec in mine:
                if not rec.get("project"):
                    continue
                g = name_match(rec["project"], pname, drop)
                if g:
                    grade = g
                    hits.append("project" if g == "strong" else "project?")
                    matched_to = rec
                    break
            if any(k in hay for k in places):
                hits.append("place")
            if any(k in dname.lower() for k in devs):
                hits.append("developer")
            in_district = dist in districts
            if in_district:
                hits.append("district")

            tu = str(p.get("unit", "")).strip()
            if tu.isdigit() and int(tu) > 0 and int(tu) in by_units:
                hits.append("units")
                if not matched_to:
                    matched_to = by_units[int(tu)]

            # What earns a place on the sheet. A name match or an Ukay place
            # name is specific enough on its own. A developer match is not:
            # Sime Darby alone has 15 registered companies building all over
            # Malaysia, so it only counts alongside something else. District
            # alone goes to the sweep rather than the main list.
            strong = (grade == "strong") or ("place" in hits)
            if not strong and len(hits) < 2:
                if in_district:
                    sweep.append((p, dev, "district only"))
                continue
            if not hits:
                continue

            teduh_units = p.get("unit")
            mine_units = (matched_to or {}).get("advertised_units", "")
            verdict = ""
            if str(teduh_units).strip().isdigit() and str(mine_units).strip().isdigit():
                verdict = "MATCH" if int(teduh_units) == int(mine_units) else \
                          f"differs by {int(teduh_units)-int(mine_units):+d}"
            rows.append((p, dev, ", ".join(hits), matched_to, mine_units, verdict))

        area_sheet(wb, area, rows)
        sweep_sheet(wb, f"{area} district sweep", sweep)


AREA_HEADS = ["MATCHED ON", "YOUR PROJECT", "PROJECT ON TEDUH", "PROJECT CODE",
              "DEVELOPER", "DEV CODE", "TEDUH UNITS", "YOUR UNITS", "UNITS CHECK",
              "DISTRICT", "STATE", "STATUS"]


def area_sheet(wb, title, rows):
    ws = wb.create_sheet(title[:31], 2)
    for c, h in enumerate(AREA_HEADS, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="EB6834") if h == "UNITS CHECK" else HEAD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # strongest evidence first: a name match beats a district coincidence
    rows.sort(key=lambda r: (-len(r[2].split(", ")), r[1].get("nama_pemaju", "") if r[1] else ""))
    for r, (p, dev, hits, mine, mine_units, verdict) in enumerate(rows, 2):
        vals = [hits, (mine or {}).get("project", ""), p.get("projek_nama", ""),
                p.get("kod_projek", ""), dev.get("nama_pemaju", ""), p.get("kod_pemaju", ""),
                int(p["unit"]) if str(p.get("unit", "")).strip().isdigit() else p.get("unit", ""),
                int(mine_units) if str(mine_units).strip().isdigit() else mine_units,
                verdict, p.get("daerah", ""), p.get("negeri", ""), p.get("status", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name=FONT, size=10)
        if verdict == "MATCH":
            ws.cell(r, 9).fill = PatternFill("solid", fgColor="D6EAD6")
        elif verdict:
            ws.cell(r, 9).fill = KNOWN
    for c, w in enumerate([26, 30, 34, 13, 38, 10, 12, 11, 15, 16, 17, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(AREA_HEADS))}{max(2, len(rows)+1)}"
    ws.row_dimensions[1].height = 28


def sweep_sheet(wb, title, rows):
    ws = wb.create_sheet(title[:31], 3)
    heads = ["PROJECT ON TEDUH", "PROJECT CODE", "DEVELOPER", "DEV CODE",
             "TEDUH UNITS", "DISTRICT", "STATE"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEAD
    for r, (p, dev, _) in enumerate(rows, 2):
        vals = [p.get("projek_nama", ""), p.get("kod_projek", ""), dev.get("nama_pemaju", ""),
                p.get("kod_pemaju", ""),
                int(p["unit"]) if str(p.get("unit", "")).strip().isdigit() else p.get("unit", ""),
                p.get("daerah", ""), p.get("negeri", "")]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v).font = Font(name=FONT, size=10)
    for c, w in enumerate([34, 13, 38, 10, 12, 16, 17], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{max(2, len(rows)+1)}"


def corroborate(dev_code, by_dev, known, by_units):
    """Does this developer own a project you already know belongs to the company?

    This is the check that scales. Matching a developer by name or address is a
    guess; finding that it owns a project on your own list, by name or by exact
    unit count, is corroboration -- and it needs no judgement from you.
    """
    for p in by_dev.get(str(dev_code), []):
        pname = p.get("projek_nama") or ""
        for rec in known:
            if rec.get("project") and name_match(rec["project"], pname):
                return f"owns {rec['project']}"
        u = str(p.get("unit", "")).strip()
        if u.isdigit() and int(u) > 0 and int(u) in by_units:
            return f"unit count {u} = {by_units[int(u)]['project']}"
    return ""


def grade_candidate(hits, corr, confirmed=False):
    """How much of your attention this row deserves."""
    if confirmed:
        return "CONFIRMED - you checked this code"
    if "email" in hits:
        return "CONFIRMED - email domain"
    if corr:
        return f"CONFIRMED - {corr}"
    if len(hits) >= 2:
        return "LIKELY - " + ", ".join(hits)
    return "CHECK - " + (", ".join(hits) or "known code")


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        ROOT, "data", "TEDUH_Company_Index.xlsx")
    devs = read(os.path.join(ROOT, "data", "developers.csv"))
    companies = read(os.path.join(ROOT, "companies.csv"))
    found = [d for d in devs if d.get("found") == "yes"]
    found.sort(key=lambda d: int(d["kod_pemaju"]) if str(d["kod_pemaju"]).isdigit() else 0)

    wb = Workbook()
    wb.remove(wb.active)
    legend(wb, devs, companies, len(found))

    ppath0 = os.path.join(ROOT, "data", "projects_index.csv")
    by_dev_all = {}
    if os.path.exists(ppath0):
        for p in read(ppath0):
            by_dev_all.setdefault(p["kod_pemaju"], []).append(p)
    cpath = os.path.join(ROOT, "company_projects.csv")
    company_projects = read(cpath) if os.path.exists(cpath) else []

    summary, needs_check, missing = [], [], []
    for rule in companies:
        known = set(parts(rule.get("known_codes")))
        known = {k.upper() for k in known}
        mine = [k for k in company_projects
                if k.get("company", "").lower() == rule["company"].lower()]
        by_units = {}
        for k in mine:
            v = str(k.get("advertised_units", "")).strip()
            if v.isdigit():
                by_units.setdefault(int(v), []).append(k)
        by_units = {u: v[0] for u, v in by_units.items() if len(v) == 1}

        cands, covered = [], set()
        for d in found:
            hits = matches(d, rule)
            if not (hits or str(d["kod_pemaju"]) in known):
                continue
            corr = corroborate(d["kod_pemaju"], by_dev_all, mine, by_units)
            row = dict(d)
            row["_matched"] = ", ".join(hits) if hits else "known code"
            row["_confidence"] = grade_candidate(
                hits, corr, str(d["kod_pemaju"]).upper() in known)
            cands.append(row)
            if corr:
                for p in by_dev_all.get(str(d["kod_pemaju"]), []):
                    pn = p.get("projek_nama") or ""
                    for rec in mine:
                        if rec.get("project") and name_match(rec["project"], pn):
                            covered.add(rec["project"])
                    u = str(p.get("unit", "")).strip()
                    if u.isdigit() and int(u) in by_units:
                        covered.add(by_units[int(u)]["project"])
            if row["_confidence"].startswith("CHECK"):
                needs_check.append((rule["company"], row))
        for rec in mine:
            if rec.get("project") and rec["project"] not in covered:
                missing.append((rule["company"], rec))
        cands.sort(key=lambda r: (r["_confidence"].startswith("CHECK"),
                                  r["_confidence"].startswith("LIKELY"),
                                  r["nama_pemaju"]))
        sheet(wb, rule["company"], cands, extra="CONFIDENCE", known=known)
        projects = sum(int(c["bilangan_projek"]) for c in cands
                       if str(c.get("bilangan_projek", "")).strip().isdigit())
        summary.append((rule["company"], len(known), len(cands), projects))

    ws = wb.create_sheet("Summary", 1)
    for c, h in enumerate(["COMPANY", "CONFIRMED CODES", "CANDIDATES", "PROJECTS IF ALL CONFIRMED"], 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEAD
    for r, row in enumerate(summary, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v).font = Font(name=FONT, size=10)
    for c, w in enumerate([24, 18, 14, 26], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # Every project belonging to a candidate developer, for checking unit
    # counts against what the company advertises.
    projects = []
    ppath = os.path.join(ROOT, "data", "projects_index.csv")
    if os.path.exists(ppath):
        by_dev = {}
        for p in read(ppath):
            by_dev.setdefault(p["kod_pemaju"], []).append(p)
        for rule in companies:
            known = {k.upper() for k in parts(rule.get("known_codes"))}
            for d in found:
                if matches(d, rule) or str(d["kod_pemaju"]) in known:
                    for p in by_dev.get(str(d["kod_pemaju"]), []):
                        row = dict(p)
                        row["_company"] = rule["company"]
                        projects.append(row)
        projects.sort(key=lambda r: (r["_company"], r["kod_pemaju"], r["kod_projek"]))

    ws = wb.create_sheet("Company projects", 2)
    heads = ["COMPANY", "DEVELOPER CODE", "DEVELOPER NAME", "PROJECT CODE",
             "PROJECT NAME ON TEDUH", "TOTAL UNITS", "ADVERTISED UNITS",
             "STATE", "DISTRICT", "STATUS", "PERMIT START", "PERMIT END"]
    keys = ["_company", "kod_pemaju", "nama_pemaju", "kod_projek", "projek_nama",
            "unit", None, "negeri", "daerah", "status", "permit_mula", "permit_tamat"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="EB6834") if h == "ADVERTISED UNITS" else HEAD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r, row in enumerate(projects, 2):
        for c, k in enumerate(keys, 1):
            if k is None:                       # left blank for you to fill in
                ws.cell(r, c).fill = PatternFill("solid", fgColor="FFF2CC")
                continue
            v = row.get(k, "")
            if k == "unit" and str(v).strip().isdigit():
                v = int(v)
            ws.cell(r, c, v).font = Font(name=FONT, size=10)
    for c, w in enumerate([18, 15, 40, 14, 38, 12, 16, 16, 16, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{max(2, len(projects) + 1)}"
    ws.row_dimensions[1].height = 28

    # The two sheets that decide where your attention goes.
    ws = wb.create_sheet("Needs checking", 1)
    heads = ["COMPANY", "WHY IT IS UNCERTAIN", "DEVELOPER CODE", "COMPANY NAME",
             "EMAIL", "BUSINESS ADDRESS", "PROJECTS", "FIRST PROJECT"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="EB6834")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r, (comp, d) in enumerate(needs_check, 2):
        vals = [comp, d.get("_confidence", ""), int(d["kod_pemaju"]) if str(d["kod_pemaju"]).isdigit() else d["kod_pemaju"],
                d.get("nama_pemaju", ""), d.get("emel", ""), d.get("alamat_perniagaan", ""),
                d.get("bilangan_projek", ""), d.get("projek_nama", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 6))
    for c, w in enumerate([18, 30, 15, 42, 30, 52, 10, 34], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{max(2, len(needs_check)+1)}"
    ws.row_dimensions[1].height = 28

    ws = wb.create_sheet("Missing", 2)
    heads = ["COMPANY", "PROJECT ON YOUR LIST", "LOCATION", "ADVERTISED UNITS",
             "NOTE"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="EB6834")
    for r, (comp, rec) in enumerate(missing, 2):
        u = str(rec.get("advertised_units", "")).strip()
        note = ("no unit count researched, so it can only be found by name"
                if not u else "")
        vals = [comp, rec.get("project", ""), rec.get("location", ""),
                int(u) if u.isdigit() else u, note]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v).font = Font(name=FONT, size=10)
    for c, w in enumerate([18, 44, 26, 16, 52], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{max(2, len(missing)+1)}"

    build_areas(wb, found, ppath)

    sheet(wb, "All developers", found)

    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"{out}: {len(found)} developers, {len(projects)} company projects, "
          f"{len(companies)} company sheets")
    print(f"  needs checking: {len(needs_check)} candidates")
    print(f"  missing:        {len(missing)} of your listed projects not covered")
    for name, k, c, p in summary:
        print(f"  {name:<20} {k} confirmed, {c} candidates, {p} projects")


if __name__ == "__main__":
    main()
