#!/usr/bin/env python3
"""Build the daily tracker workbook from data/teduh_daily.csv.

One sheet per tracker, one column group (NEW / SOLD / %) per day the scraper has
run, newest day first. This is the day-by-day companion to the weekly Excel
files, which stay on their Friday rhythm.

Usage: python3 scripts/build_daily_xlsx.py <projects.csv> <teduh_daily.csv> <output.xlsx>
"""
import csv, os, sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter as L

# The four area sheets, in the order they should appear in the workbook.
# Everything else in projects.csv is a developer tracker and gets a sheet of the
# same shape, named from its tracker_label column -- so a new competitor needs
# rows in projects.csv and nothing here.
AREA_SHEET = {"seputeh": "Seputeh Hills", "status13": "Klang Valley",
              "johor": "Johor", "ukay": "Ukay"}


def sheet_order(projects):
    """[(tracker key, sheet title)] -- areas first, then developers A-Z."""
    labels, areas, devs = {}, [], []
    for p in projects:
        key = (p.get("tracker") or "").strip()
        if not key or key in labels:
            continue
        if key in AREA_SHEET:
            labels[key] = AREA_SHEET[key]
            areas.append(key)
        else:
            labels[key] = (p.get("tracker_label") or "").strip() or key.title()
            devs.append(key)
    areas.sort(key=lambda k: list(AREA_SHEET).index(k))
    devs.sort(key=lambda k: labels[k].lower())
    return [(k, labels[k]) for k in areas + devs]


def F(**k):
    k.setdefault("size", 10)
    return Font(name="Arial", **k)


TH = Side(style="thin")
BORD = Border(left=TH, right=TH, top=TH, bottom=TH)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
HDR = PatternFill("solid", fgColor="D9E1F2")
NEWF = PatternFill("solid", fgColor="FCE4D6")


def cell(ws, r, c, v=None, font=None, align=CTR, fill=None, fmt=None):
    x = ws.cell(r, c, v)
    x.font = font or F()
    x.alignment = align
    x.border = BORD
    if fill:
        x.fill = fill
    if fmt:
        x.number_format = fmt
    return x


def main():
    pcsv, dcsv, out = sys.argv[1], sys.argv[2], sys.argv[3]
    projects = list(csv.DictReader(open(pcsv, encoding="utf-8")))
    rows = list(csv.DictReader(open(dcsv, encoding="utf-8")))

    lookup, days = {}, set()
    for r in rows:
        try:
            v = int(float(r["total_sold"]))
        except (TypeError, ValueError):
            continue
        lookup[(r["code"], r["week"])] = v
        days.add(r["week"])
    days = sorted(days, reverse=True)          # newest first
    chrono = list(reversed(days))

    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.remove(wb.active)

    for key, title in sheet_order(projects):
        mine = sorted((p for p in projects if p["tracker"] == key),
                      key=lambda x: (x.get("pin", "").strip().lower() not in ("yes", "y", "1", "true")))
        if not mine:
            continue
        ws = wb.create_sheet(title[:31])
        ws["A1"] = title + " — daily tracker"
        ws["A1"].font = F(bold=True, size=12)
        ws["A2"] = "Every day the tracker has run, newest first. Generated " + datetime.now().strftime("%d/%m/%Y")
        ws["A2"].font = F(bold=True, color="C00000")

        for i, h in enumerate(["NO", "PROJECT", "APDL \nDATE", "TOTAL \nUNITS", "PROJECT \nCODE", "DEVELOPER"], 1):
            cell(ws, 4, i, h, F(bold=True), fill=HDR)
        col = 7
        for i, d in enumerate(days):
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
            cell(ws, 3, col, datetime.strptime(d, "%Y-%m-%d"), F(bold=True), fmt="DD/MM/YYYY",
                 fill=NEWF if i == 0 else None)
            for j, h in enumerate(["NEW SALES", "TOTAL SOLD", "%"]):
                cell(ws, 4, col + j, h, F(bold=True), fill=NEWF if i == 0 else HDR)
            col += 3

        for r, p in enumerate(mine, 5):
            code = (p.get("code") or "").split(",")[0].strip()
            units = int(p["total_units"]) if str(p.get("total_units", "")).strip().isdigit() else 0
            cell(ws, r, 1, int(p["no"]))
            cell(ws, r, 2, (p["project"] or "").replace("\n", " ").strip(), align=LFT)
            apdl = (p.get("apdl") or "").strip()
            cell(ws, r, 3, datetime.strptime(apdl, "%Y-%m-%d") if apdl else None,
                 fmt="DD/MM/YYYY")
            cell(ws, r, 4, units, fmt="#,##0")
            cell(ws, r, 5, code or "-")
            cell(ws, r, 6, p["developer"], align=LFT)

            # deltas computed chronologically, then written newest-first
            delta, prev = {}, None
            for d in chrono:
                v = lookup.get((code, d))
                if v is None:
                    continue
                if prev is not None:
                    delta[d] = v - prev
                prev = v

            col = 7
            for i, d in enumerate(days):
                fill = NEWF if i == 0 else None
                v = lookup.get((code, d))
                cell(ws, r, col, delta.get(d), fmt="#,##0", fill=fill)
                cell(ws, r, col + 1, v, fmt="#,##0", fill=fill)
                pc = cell(ws, r, col + 2, None, fmt="0.0%", fill=fill)
                if v is not None and units:
                    pc.value = f"={L(col + 1)}{r}/$D${r}"
                col += 3

        # C is 13, not 12: DD/MM/YYYY plus the cell's own padding needs it, and
        # at 12 Excel renders the whole APDL column as ##### instead.
        for c, w in [("A", 4), ("B", 26), ("C", 13), ("D", 8), ("E", 10), ("F", 30)]:
            ws.column_dimensions[c].width = w
        for c in range(7, col):
            ws.column_dimensions[L(c)].width = 9
        ws.freeze_panes = "G5"
        ws.row_dimensions[4].height = 30

    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
    wb.save(out)
    print(f"{out}  ({len(days)} daily snapshots)")


if __name__ == "__main__":
    main()
